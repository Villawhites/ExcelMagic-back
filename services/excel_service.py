import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import os
from datetime import datetime

class ExcelService:
    def __init__(self, output_dir="generated_files"):
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def create_excel(self, structure: dict, filename_prefix: str = "template") -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        columns = structure.get("columns", [])
        formulas = structure.get("formulas", {})
        styles = structure.get("styles", {})

        # Set Headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

            # Apply formulas to the first data row (row 2) if exists
            if col_name in formulas:
                ws.cell(row=2, column=col_idx, value=formulas[col_name])

        # Auto-adjust columns width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath

excel_service = ExcelService()
